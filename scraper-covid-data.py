import requests
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
out_list = []
# Collect the github page
page = requests.get('https://www.mass.gov/info-details/archive-of-covid-19-cases-in-massachusetts')
# Create a BeautifulSoup object
soup = BeautifulSoup(page.text, 'html.parser')
repo = soup.find(class_="main-content main-content--two")
repo_list = repo.find_all('a')
#month = input("Enter month:").lower()
#day = input("Enter day:")
#year = input("Enter year:")
year = str(datetime.date.today().year)
months = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']
month = months[datetime.date.today().month-1]
day = str(datetime.date.today().day+1)#tomorrow
found = 0
arr = []
while found == 0:
    day = str(int(day) - 1)
    for r in repo_list:
        arr = r.get('href').split(month+'-'+day+'-'+year)
        if(len(arr) >1):
            found  = 1
            url = r.get('href')
            url = 'https://www.mass.gov' + url
            res = requests.get(url)
            with open('./data'+month+day+year+'.xlsx', 'wb') as f:
                f.write(res.content)

            wb  = load_workbook(filename='data'+month+day+year+'.xlsx')
            ws = wb['DeathsReported (Report Date)']     
            covid_dict  = {}
            column_headers = {}      
            first_row = ws[1]
            for cell in first_row[1:]:
                column_headers[cell.column] = cell.internal_value
            for row in ws.iter_rows(min_row=2):
                #store the date out_list
                date_cell = row[1]
                temp_cell_dict = {}
                for  cell in row[1:]:   
                    col_name = column_headers[cell.column] 
                    temp_cell_dict[col_name] = cell.internal_value
                    #set date key to be contents of rows
                    covid_dict[date_cell.internal_value] = temp_cell_dict   
            
            data_json = json.dumps(covid_dict)
            json_object = json.loads(data_json)
            json_formatted_str = json.dumps(json_object, indent=2)
            with open('new'+month+day+year+'.json', "w") as f:
                f.write(json_formatted_str)